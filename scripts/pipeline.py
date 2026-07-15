#!/usr/bin/env python3
"""Merge, dedupe, verify, score, and export leads to CSV + Excel + report.

Accepts one or more JSON files (Apify Google Maps format, optionally enriched).
Stdlib only; Excel export uses openpyxl if installed (falls back to CSV-only).

  python3 pipeline.py raw1.json raw2.json --query "web agencies Hamburg" --output-dir output/
"""
import argparse
import csv
import json
import os
import re
import socket
import sys
import time
from datetime import date

EMAIL_RE = re.compile(r"^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$")
JUNK_EMAIL_RE = re.compile(r"\.(png|jpg|jpeg|gif|svg|webp|css|js)$|@(example|sentry|wixpress|domain)\.",
                           re.IGNORECASE)

COLUMNS = ["score", "icp_score", "founder_name", "founder_role", "founder_linkedin",
           "name", "category", "phone", "email", "email_status", "website",
           "linkedin", "facebook", "instagram", "x_twitter", "youtube", "tiktok",
           "full_address", "city", "postal_code", "country", "rating", "reviews_count",
           "nd_legal_name", "nd_legal_form", "nd_status", "nd_register_id", "register_status",
           "google_maps_url", "place_id", "flags", "icp_reason"]

SOCIAL_HOSTS = [("linkedin.com/in/", "founder_linkedin"), ("linkedin.com", "linkedin"),
                ("facebook.com", "facebook"), ("instagram.com", "instagram"),
                ("twitter.com", "x_twitter"), ("x.com/", "x_twitter"),
                ("youtube.com", "youtube"), ("tiktok.com", "tiktok")]


def split_socials(urls):
    """One column per platform (first URL wins); personal LinkedIn goes to founder_linkedin."""
    out = {"linkedin": "", "facebook": "", "instagram": "", "x_twitter": "",
           "youtube": "", "tiktok": "", "founder_linkedin": ""}
    for u in urls:
        ul = u.lower()
        for marker, col in SOCIAL_HOSTS:
            if marker in ul:
                if not out[col]:
                    out[col] = u
                break
    return out


def norm_phone(phone):
    if not phone:
        return ""
    digits = re.sub(r"[^\d+]", "", str(phone))
    return digits


def norm_domain(url):
    if not url:
        return ""
    url = str(url).strip().lower()
    url = re.sub(r"^https?://", "", url)
    url = re.sub(r"^www\.", "", url)
    return url.split("/")[0]


def domain_resolves(domain, cache={}):
    """DNS check with retries — flaky local resolvers (observed: ~37%% transient SERVFAIL)
    must not condemn live domains. Only 3 consecutive failures count as dead."""
    if not domain:
        return False
    if domain in cache:
        return cache[domain]
    for attempt in range(3):
        try:
            socket.getaddrinfo(domain, 443, proto=socket.IPPROTO_TCP)
            cache[domain] = True
            return True
        except OSError:
            if attempt < 2:
                time.sleep(1.5 * (attempt + 1))
    cache[domain] = False
    return False


def parse_address(addr):
    """Best-effort city/postal/country from a flat address string.

    Handles 'Straße 1, 66450 Bexbach, Germany' (EU) and
    '123 Main St, Minneapolis, MN 55401, United States' (US) shapes.
    """
    addr = str(addr or "")
    segs = [s.strip() for s in addr.split(",") if s.strip()]
    postal, city, country = "", "", ""
    m = re.search(r"\b(\d{4,6}(?:-\d{4})?)\b", addr)
    if m:
        postal = m.group(1)
    if len(segs) >= 3 and not re.search(r"\d", segs[-1]):
        country = segs[-1]
        segs = segs[:-1]
    for i, s in enumerate(segs):
        if postal and postal in s:
            rest = re.sub(r"\b[A-Z]{2}\b\s*$", "", s.replace(postal, "").strip(" -")).strip()
            city = rest or (segs[i - 1] if i > 0 else "")
            break
    if not city and len(segs) >= 2:
        city = re.sub(r"\b\d{4,6}\b|\b[A-Z]{2}\b\s*$", "", segs[-1]).strip()
    return city, postal, country


def normalize(item):
    """Map a raw Apify (or North Data power-search) item to a flat lead dict."""
    nd = item.get("northdata") or {}
    emails = item.get("emails") or ([item["email"]] if item.get("email") else [])
    if isinstance(emails, str):
        emails = [emails]
    socials = item.get("socialLinks") or []
    for k in ("instagrams", "facebooks", "linkedIns", "twitters", "youtubes", "tiktoks"):
        v = item.get(k)
        if isinstance(v, list):
            socials.extend(v)
    # North Data power-search shape (name/address dicts) vs Apify flat shape
    if "title" not in item and isinstance(item.get("name"), dict):
        addr = item.get("address") or {}
        return {
            "name": item["name"].get("name", ""), "category": item.get("segment", ""),
            "website": "", "emails": [], "phone": "",
            "full_address": ", ".join(filter(None, [addr.get("street"), addr.get("postalCode"), addr.get("city")])),
            "city": addr.get("city", ""), "postal_code": addr.get("postalCode", ""),
            "country": addr.get("country", ""), "rating": "", "reviews_count": "",
            "social_links": [], "google_maps_url": "", "place_id": "",
            "nd_legal_name": item["name"].get("name", ""), "nd_legal_form": item["name"].get("legalForm", ""),
            "nd_status": item.get("status", ""), "nd_register_id": (item.get("register") or {}).get("id", ""),
        }
    lead = {
        "name": item.get("title") or item.get("name") or "",
        "category": item.get("categoryName", ""),
        "website": item.get("website") or "",
        "emails": [e for e in emails if e],
        "email_status": item.get("email_status", ""),
        "phone": item.get("phone") or item.get("phoneUnformatted") or "",
        "full_address": item.get("address", ""),
        "city": item.get("city", ""),
        "postal_code": item.get("postalCode", ""),
        "country": item.get("countryCode", ""),
        "rating": item.get("totalScore", ""),
        "reviews_count": item.get("reviewsCount", ""),
        **split_socials(sorted(set(socials))),
        "founder_name": item.get("founder_name", ""),
        "founder_role": item.get("founder_role", ""),
        "register_status": item.get("register_status", ""),
        "icp_score": item.get("icp_score", ""),
        "icp_reason": item.get("icp_reason", ""),
        "_carry_flags": item.get("_flags") or [],
        "google_maps_url": item.get("url", ""),
        "place_id": item.get("placeId", ""),
        "nd_legal_name": nd.get("legalName", ""),
        "nd_legal_form": nd.get("legalForm", ""),
        "nd_status": nd.get("status", ""),
        "nd_register_id": nd.get("registerId", ""),
    }
    if not (lead["city"] and lead["postal_code"]):
        c, p, co = parse_address(lead["full_address"])
        lead["city"] = lead["city"] or c
        lead["postal_code"] = lead["postal_code"] or p
        lead["country"] = lead["country"] or co
    if item.get("founder_linkedin"):  # explicit find beats URL-pattern guess
        lead["founder_linkedin"] = item["founder_linkedin"]
    return lead


def score(lead, flags):
    # Phone-first weighting: German UWG restricts cold email, so calls come first.
    s = 0
    s += 25 if lead["phone"] else 0
    s += 20 if lead["email"] else 0
    s += 15 if lead["website"] else 0
    s += 15 if lead.get("founder_name") else 0
    s += 5 if any(lead.get(k) for k in ("linkedin", "facebook", "instagram", "x_twitter",
                                        "youtube", "tiktok", "founder_linkedin")) else 0
    s += 10 if (lead["nd_register_id"] or lead.get("register_status") == "registered") else 0
    s += 10 if lead.get("email_status") == "valid" else 0
    s -= 20 if lead.get("email_status") == "invalid" else 0
    try:
        r, c = float(lead["rating"] or 0), int(lead["reviews_count"] or 0)
        s += min(8, int(r * 1.6)) if c >= 3 else 0
        s += 2 if c >= 25 else 0
    except (TypeError, ValueError):
        pass
    s -= 15 * sum(1 for f in flags if f.startswith("dead_") or f in
                  ("invalid_email", "placeholder_website", "register_id_mismatch"))
    return max(0, min(100, s))


def main():
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("inputs", nargs="+", help="One or more JSON lead files to merge")
    p.add_argument("--query", default="leads", help="Human description used in filenames/report")
    p.add_argument("--output-dir", default="output")
    p.add_argument("--no-dns", action="store_true", help="Skip DNS verification (offline mode)")
    p.add_argument("--min-score", type=int, default=0, help="Drop leads scoring below this")
    p.add_argument("--require-founder", action="store_true",
                   help="Keep only leads with an identified founder/owner (run founders.py first)")
    p.add_argument("--icp-threshold", type=int, default=None,
                   help="Keep only leads with icp_score >= N (unscored leads are excluded too)")
    p.add_argument("--strict-quality", action="store_true",
                   help="Drop leads with placeholder/parked or dead websites")
    p.add_argument("--include-category", action="append",
                   help="Keep only leads whose category contains one of these substrings "
                        "(case-insensitive, repeatable) — guards against Google Maps sweeping "
                        "in adjacent categories (gyms for 'wellness coach' etc.)")
    p.add_argument("--exclude-category", action="append",
                   help="Drop leads whose category contains one of these substrings (repeatable)")
    args = p.parse_args()

    raw = []
    for path in args.inputs:
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
            raw.extend(data if isinstance(data, list) else [data])
    print(f"Loaded {len(raw)} raw records from {len(args.inputs)} file(s).")

    # --- dedupe ---
    seen_pid, seen_phone, seen_domain, seen_name = set(), set(), set(), set()
    leads, dupes = [], 0
    for item in raw:
        lead = normalize(item)
        if not lead["name"]:
            continue
        pid = lead["place_id"]
        phone = norm_phone(lead["phone"])
        domain = norm_domain(lead["website"])
        namekey = re.sub(r"\W+", "", lead["name"].lower()) + "|" + str(lead.get("postal_code", ""))
        if (pid and pid in seen_pid) or (phone and phone in seen_phone) or \
           (domain and domain in seen_domain) or (not pid and namekey in seen_name):
            dupes += 1
            continue
        if pid:
            seen_pid.add(pid)
        if phone:
            seen_phone.add(phone)
        if domain:
            seen_domain.add(domain)
        seen_name.add(namekey)
        lead["_domain"] = domain
        lead["phone"] = phone
        leads.append(lead)
    print(f"After dedupe: {len(leads)} unique leads ({dupes} duplicates removed).")

    # --- verify ---
    flagged = 0
    for lead in leads:
        flags = list(lead.pop("_carry_flags", []))
        emails = [e for e in lead.pop("emails", [])
                  if EMAIL_RE.match(e) and not JUNK_EMAIL_RE.search(e)]
        if lead.get("_domain"):
            own = [e for e in emails if e.split("@")[-1].endswith(lead["_domain"])]
            emails = own + [e for e in emails if e not in own]
        lead["email"] = emails[0] if emails else ""
        if not emails and lead.get("_email_enriched"):
            flags.append("no_email_found")
        if not args.no_dns:
            dead_site = lead["_domain"] and not domain_resolves(lead["_domain"])
            if dead_site:
                flags.append("dead_website_domain")
            edom = lead["email"].split("@")[-1] if lead["email"] else ""
            if edom and ((edom == lead["_domain"] and dead_site) or
                         (edom != lead["_domain"] and not domain_resolves(edom))):
                flags.append("invalid_email")
                lead["email"] = ""
        if not lead["website"] and not lead["phone"]:
            flags.append("no_contact_channel")
        lead["flags"] = ";".join(flags)
        lead["score"] = score(lead, flags)
        if flags:
            flagged += 1
        lead.pop("_domain", None)

    # --- qualification filters ---
    dropped = {}
    if args.include_category:
        n0 = len(leads)
        pats = [c.lower() for c in args.include_category]
        leads = [l for l in leads if any(p in (l.get("category") or "").lower() for p in pats)]
        dropped["category_not_matching"] = n0 - len(leads)
    if args.exclude_category:
        n0 = len(leads)
        pats = [c.lower() for c in args.exclude_category]
        leads = [l for l in leads if not any(p in (l.get("category") or "").lower() for p in pats)]
        dropped["category_excluded"] = n0 - len(leads)
    if args.require_founder:
        n0 = len(leads)
        leads = [l for l in leads if l.get("founder_name")]
        dropped["no_founder_identified"] = n0 - len(leads)
    if args.strict_quality:
        n0 = len(leads)
        leads = [l for l in leads if "placeholder_website" not in l["flags"]
                 and "dead_website_domain" not in l["flags"]]
        dropped["placeholder_or_dead_website"] = n0 - len(leads)
    if args.icp_threshold is not None:
        n0 = len(leads)
        kept = []
        for l in leads:
            try:
                if int(l.get("icp_score") or -1) >= args.icp_threshold:
                    kept.append(l)
            except (TypeError, ValueError):
                pass  # unscoreable value -> excluded
        leads = kept
        dropped[f"icp_below_{args.icp_threshold}_or_unscored"] = n0 - len(leads)
    for k, v in dropped.items():
        if v:
            print(f"Filtered out {v} leads: {k}")

    leads = [l for l in leads if l["score"] >= args.min_score]
    leads.sort(key=lambda l: (-int(l["icp_score"]) if str(l.get("icp_score", "")).isdigit() else 0,
                              -l["score"]))

    # --- export ---
    os.makedirs(args.output_dir, exist_ok=True)
    slug = re.sub(r"\W+", "_", args.query.lower()).strip("_")[:50] or "leads"
    csv_path = os.path.join(args.output_dir, f"{slug}.csv")
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=COLUMNS, extrasaction="ignore")
        w.writeheader()
        w.writerows(leads)
    print(f"CSV → {csv_path}")

    xlsx_path = os.path.join(args.output_dir, f"{slug}.xlsx")
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
        wb = Workbook()
        ws = wb.active
        ws.title = "Leads"
        ws.append([c.replace("_", " ").title() for c in COLUMNS])
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E79")
        for lead in leads:
            ws.append([lead.get(c, "") for c in COLUMNS])
        widths = {"name": 32, "founder_name": 24, "founder_role": 16, "founder_linkedin": 30,
                  "website": 28, "email": 28, "phone": 18, "full_address": 38, "category": 22,
                  "linkedin": 26, "facebook": 26, "instagram": 26, "x_twitter": 22,
                  "youtube": 22, "tiktok": 22, "google_maps_url": 18, "flags": 22,
                  "icp_reason": 44}
        for i, c in enumerate(COLUMNS, 1):
            ws.column_dimensions[get_column_letter(i)].width = widths.get(c, 12)
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"
        wb.save(xlsx_path)
        print(f"Excel → {xlsx_path}")
    except ImportError:
        xlsx_path = None
        print("openpyxl not installed — skipped Excel export (pip install openpyxl).")

    # --- report ---
    n = len(leads)
    def cnt(k):
        return sum(1 for l in leads if l.get(k))
    report_path = os.path.join(args.output_dir, "report.md")
    with open(report_path, "w", encoding="utf-8") as f:
        f.write(f"# Lead report: {args.query}\n\nGenerated {date.today().isoformat()}\n\n")
        f.write(f"- Raw records: {len(raw)} from {len(args.inputs)} file(s)\n")
        f.write(f"- Duplicates removed: {dupes}\n")
        f.write(f"- **Final leads: {n}**\n")
        for k, v in dropped.items():
            f.write(f"- Filtered out ({k}): {v}\n")
        if n:
            f.write(f"- With phone: {cnt('phone')} ({100*cnt('phone')//n}%)\n")
            f.write(f"- With email: {cnt('email')} ({100*cnt('email')//n}%)\n")
            f.write(f"- With website: {cnt('website')} ({100*cnt('website')//n}%)\n")
            f.write(f"- Founder identified: {cnt('founder_name')} ({100*cnt('founder_name')//n}%)\n")
            f.write(f"- Registry-verified: {cnt('nd_register_id')} ND / "
                    f"{sum(1 for l in leads if l.get('register_status') == 'registered')} Handelsregister\n")
            f.write(f"- Flagged for review: {flagged}\n")
            top_flags = {}
            for l in leads:
                for fl in filter(None, l["flags"].split(";")):
                    top_flags[fl] = top_flags.get(fl, 0) + 1
            if top_flags:
                f.write("\n## Flags\n\n")
                for fl, c in sorted(top_flags.items(), key=lambda x: -x[1]):
                    f.write(f"- `{fl}`: {c}\n")
    print(f"Report → {report_path}")
    print(f"\nSummary: {n} leads | email {100*cnt('email')//n if n else 0}% | "
          f"website {100*cnt('website')//n if n else 0}% | flagged {flagged}")


if __name__ == "__main__":
    main()
